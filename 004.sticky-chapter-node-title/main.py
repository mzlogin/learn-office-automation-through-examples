import openpyxl

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.active

chapterName = None
nodeName = None

for row in sheet.iter_rows():
    ifNode = False
    if row[0].value != None and row[0].value != '' and row[0].value != chapterName:
        chapterName = row[0].value
    if row[1].value != None and row[1].value != '' and row[1].value != nodeName:
        nodeName = row[1].value
        ifNode = True
    if ifNode:
        if row[2].value is not None:
            sheet.cell(row=row[0].row, column=4).value = chapterName + nodeName + ' ' + row[2].value
    else:
        sheet.cell(row=row[0].row, column=5).value = chapterName + nodeName + row[2].value
wb.save('test.xlsx')